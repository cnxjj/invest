import pandas as pd
import efinance as ef
import openpyxl as op
from datetime import datetime
from openpyxl import load_workbook

# 遍历excel文件内的所有sheet，返回列表
def get_sheet_names(file):
    wb = op.load_workbook(file) # 获取workbook中所有的表格
    ws = wb.sheetnames # 循环遍历所有sheet
    return(ws)

# 判断给定日期是周几，返回0-6
def isweek(date):
    date = datetime.strptime(end, "%Y%m%d") # 字符串转为日期格式
    week = datetime.weekday(date)
    return(week)

# 基金代码列表
fund_codes = ['000300','162412','110031','006327','000961','012348','501029','003318']
fund_names = ['沪深300','华宝中证医疗', '易方达恒生国企','易方达中概互联','天弘沪深300','天弘恒生科技','华宝红利基金','景顺中证500低波动']

# 股票代码列表
stock_codes = ['春秋航空', '洋河股份', '分众传媒', '腾讯控股', '小米集团', '福寿园']

beg = '20230801'  # 开始日期
end = datetime.today().strftime("%Y%m%d")  # 取当前日期转换为20230822格式
file = r'20_Invest/stock/'+end+'.xlsx'
writer = pd.ExcelWriter(file)
count_fund = 0  # 基金编号计数器
count_stock = 0  # 股票编号计数器

# 获取多只基金数据，存储为xlsx中的多个sheet
#for fund_code in fund_codes:
#    df_fund = ef.fund.get_quote_history(fund_code)
#    df_fund.to_excel(writer, sheet_name=f'{count_fund} - {fund_names[count_fund]}', index=False)
#    print(f'基金: {count_fund} - {fund_code} 的行情数据已存储到文件中')
#    count_fund = count_fund + 1
#print('全部基金数据获取完毕！')

# 获取多只股票数据，存储为xlsx中的多个sheet
for stock_code in stock_codes:
    df_stock = ef.stock.get_quote_history(stock_code, beg=beg, end=end)
    df_stock.sort_index(ascending=False,inplace=True) # 按时间倒序
    df_stock.to_excel(writer, sheet_name=f'{count_stock + 8} - {stock_code}', index=False)
    print(f'股票: {count_stock + 8} - {stock_code} 的行情数据已存储到文件中')
    count_stock = count_stock + 1
print('全部股票获取完毕！')

writer._save()
writer.close()

# 打开生成的xlsx文件进行汇总处理
wb = load_workbook(file)
ws = wb.create_sheet(index=0, title='汇总') # 添加汇总工作簿

for sheet in wb:
    print(sheet.title)


wb.save(file)
print('汇总完毕！')