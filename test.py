import openpyxl as op
import pandas as pd
import xlwings as xw
from openpyxl import load_workbook
from datetime import datetime

# 判断给定日期是周几，返回0-6
def isweek(date):
    date_f = datetime.strptime(date, '%Y%m%d') # 字符串转为日期格式
    week = datetime.weekday(date_f)
    # print(date_f, week)
    return(week)

def copy2cell(ws_src, ws_dsc, src, dsc):
    print(src)
    ws_dsc[dsc].value = str(src)  # 修改单元格的值

    #cell_src = ws_src[src]
    #cell_dsc = ws_dsc[dsc]
    # cell_dsc.value = str(cell_src.value)  # 修改单元格的值
    # print('      ', ws_dsc, + '=====>' + cell_dsc.value)


# 将每周六的收盘收据汇总
def move_date(file_name, src_sheet, dsc_sheet):
    wb = load_workbook(file_name)
    ws_src = wb[src_sheet]
    ws_dsc = wb[dsc_sheet]

    for row in ws_src.iter_rows():
        row[2].value = str(row[2].value.replace('-',''))
        if row[2].value != '日期':
            if isweek(row[2].value) == 4:
                # print(row[2].value, row[4].value)
                copy2cell(ws_src, ws_dsc, row[4].value, 'B2')
        
beg = '20230901'  # 开始日期
end = datetime.today().strftime("%Y%m%d")  # 取当前日期转换为20230822格式
file = r'20_Invest/stock/'+end+'.xlsx'

move_date(file, '8 - 春秋航空', '汇总')